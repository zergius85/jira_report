# -*- coding: utf-8 -*-
"""
Jira Report System — Ядро отчётов

Модуль для сбора, обработки и выгрузки данных из Jira.
Поддерживает консольный режим и работу через Web-интерфейс.
"""
import re
from typing import Optional, List, Dict, Any, Tuple, Union, Callable
from jira import JIRA, JIRAError
import pandas as pd
import warnings
from urllib3.exceptions import InsecureRequestWarning
from datetime import datetime, timedelta
import argparse
import os
import sys
import logging
import io
from dateutil.relativedelta import relativedelta
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type
import requests
from requests.auth import HTTPBasicAuth

# Импортируем настройки из core.config
from core.config import (
    JIRA_SERVER,
    JIRA_USER,
    JIRA_PASS,
    EXCLUDED_PROJECTS,
    INTERNAL_PROJECTS,
    CLOSED_STATUS_IDS,
    EXCLUDED_ASSIGNEE_CLOSE,
    EXCLUDED_PROJECTS_NO_TIMESPENT,
    SSL_VERIFY,
    REPORT_BLOCKS,
    LOG_LEVEL,
    LOG_FORMAT,
    LOG_DATE_FORMAT,
    MAX_REPORT_DAYS,
    RISK_ZONE_INACTIVITY_THRESHOLD,
    MAX_SEARCH_RESULTS,
    MAX_EXCEL_ROWS
)

# Импортируем утилиты
from core.utils import sanitize_jql_identifier, sanitize_jql_string_literal

# --- НАСТРОЙКА ЛОГИРОВАНИЯ ---
logging.basicConfig(
    level=getattr(logging, LOG_LEVEL),
    format=LOG_FORMAT,
    datefmt=LOG_DATE_FORMAT
)
logger = logging.getLogger(__name__)

# --- НАСТРОЙКА SSL ---
if not SSL_VERIFY:
    logger.warning("⚠️  Проверка SSL отключена (SSL_VERIFY=false)")
    # SSL отключается только для подключения к Jira через options={'verify': False}
    # Глобальные переменные окружения НЕ изменяются для безопасности
    warnings.simplefilter('ignore', InsecureRequestWarning)
else:
    logger.info("✅ Проверка SSL включена")

def validate_config() -> Tuple[bool, List[str]]:
    """
    Проверяет корректность конфигурации.
    
    Returns:
        Tuple[bool, List[str]]: (успех, список ошибок)
    """
    errors = []
    
    if not JIRA_SERVER:
        errors.append("Не указан JIRA_SERVER в .env")
    if not JIRA_USER:
        errors.append("Не указан JIRA_USER в .env")
    if not JIRA_PASS:
        errors.append("Не указан JIRA_PASS в .env")
    
    return (len(errors) == 0, errors)


@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=1, min=2, max=10),
    retry=retry_if_exception_type((ConnectionError, JIRAError)),
    reraise=True
)
def get_jira_connection() -> JIRA:
    """
    Устанавливает соединение с Jira с автоматическими повторными попытками.

    Returns:
        JIRA: Объект подключения к Jira

    Raises:
        ConnectionError: При ошибке подключения после всех попыток
    """
    try:
        logger.info(f"🔌 Подключение к Jira: {JIRA_SERVER}")

        if not SSL_VERIFY:
            jira = JIRA(
                server=JIRA_SERVER,
                basic_auth=(JIRA_USER, JIRA_PASS),
                options={'verify': False}
            )
        else:
            jira = JIRA(
                server=JIRA_SERVER,
                basic_auth=(JIRA_USER, JIRA_PASS)
            )

        # Проверка подключения
        jira.myself()
        logger.info("✅ Успешное подключение к Jira")
        return jira

    except JIRAError as e:
        logger.error(f"❌ Ошибка подключения к Jira: {e.text}")
        raise ConnectionError(f"Не удалось подключиться к Jira: {e.text}")
    except Exception as e:
        logger.error(f"❌ Неизвестная ошибка подключения: {e}")
        raise ConnectionError(f"Ошибка подключения: {e}")

def get_default_start_date() -> datetime:
    """
    Возвращает дату начала по умолчанию (1 число прошлого месяца).
    
    Returns:
        datetime: Дата начала
    """
    today = datetime.now()
    if today.month == 1:
        return datetime(today.year - 1, 12, 1)
    else:
        return datetime(today.year, today.month - 1, 1)


def convert_seconds_to_hours(seconds: Optional[int]) -> float:
    """
    Конвертирует секунды в часы.

    Args:
        seconds: Время в секундах

    Returns:
        float: Время в часах
    """
    if seconds is None:
        return 0.0
    return round(seconds / 3600, 2)


def normalize_filter(
    value: Optional[Union[str, List[str]]],
    upper: bool = False
) -> List[str]:
    """
    Нормализует значение фильтра: строку превращает в список, None в пустой список.

    Args:
        value: Значение фильтра (строка, список или None)
        upper: Преобразовать ли в верхний регистр

    Returns:
        List[str]: Нормализованный список значений
    """
    if isinstance(value, str):
        value = value.upper() if upper else value
        return [value]
    elif value is None:
        return []
    else:
        return [v.upper() if upper else v for v in value]


def search_all_issues(
    jira: JIRA,
    jql: str,
    fields: str = '*all',
    expand: Optional[str] = None,
    batch_size: int = 100
) -> List[Any]:
    """
    Выполняет поиск задач с пагинацией для получения более 5000 результатов.

    Jira API ограничивает максимальное количество результатов до 5000 за запрос.
    Эта функция разбивает запрос на батчи по batch_size задач.

    Args:
        jira: Подключение к Jira
        jql: JQL-запрос
        fields: Поля для получения (по умолчанию '*all')
        expand: Дополнительные данные (например, 'changelog')
        batch_size: Размер батча (по умолчанию 100)

    Returns:
        List[Any]: Список всех задач
    """
    all_issues: List[Any] = []
    start_at = 0

    logger.info(f"Выполнение поиска с пагинацией: {jql[:100]}...")
    
    while True:
        batch = jira.search_issues(
            jql,
            startAt=start_at,
            maxResults=batch_size,
            fields=fields,
            expand=expand
        )
        
        if not batch:
            break
            
        all_issues.extend(batch)
        logger.debug(f"Получено {len(batch)} задач (всего: {len(all_issues)})")
        
        # Если получили меньше, чем запросили — это последняя страница
        if len(batch) < batch_size:
            break
            
        start_at += batch_size
    
    logger.info(f"Поиск завершён. Всего получено задач: {len(all_issues)}")
    return all_issues


def fetch_issues_via_rest(
    jira: JIRA,
    jql: str,
    batch_size: int = 100
) -> List[Dict[str, Any]]:
    """
    Получает задачи через REST API с полем creator.
    
    Jira Python client не возвращает creator корректно, поэтому используем
    прямой REST API запрос с expand=names,schema.
    
    Args:
        jira: Подключение к Jira
        jql: JQL-запрос
        batch_size: Размер батча
        
    Returns:
        List[Dict]: Список задач с полями
    """
    all_issues = []
    start_at = 0
    
    logger.info(f"REST API запрос: {jql[:100]}...")
    
    # Используем сессию для аутентификации
    session = requests.Session()
    session.auth = HTTPBasicAuth(JIRA_USER, JIRA_PASS)
    if not SSL_VERIFY:
        urllib3 = requests.packages.urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    
    while True:
        url = f"{JIRA_SERVER}/rest/api/2/search"
        params = {
            'jql': jql,
            'startAt': start_at,
            'maxResults': batch_size,
            'fields': 'summary,assignee,timespent,timeoriginalestimate,resolutiondate,issuetype,duedate,status,created,updated,creator,priority,project',
            'expand': 'names'  # Для получения имён полей
        }
        
        response = session.get(url, params=params, verify=SSL_VERIFY)
        response.raise_for_status()
        
        data = response.json()
        issues = data.get('issues', [])
        
        if not issues:
            break
        
        all_issues.extend(issues)
        logger.debug(f"Получено {len(issues)} задач (всего: {len(all_issues)})")
        
        if len(issues) < batch_size:
            break
            
        start_at += batch_size
    
    logger.info(f"REST API запрос завершён. Всего получено задач: {len(all_issues)}")
    return all_issues


def get_closed_status_ids() -> List[str]:
    """
    Автоматически определяет ID статуса "Закрыт" в Jira.
    Кэширует результат в .env для последующих запусков.
    
    Returns:
        List[str]: Список ID статусов
    """
    if CLOSED_STATUS_IDS and CLOSED_STATUS_IDS[0] != '':
        logger.info(f"✅ ID статуса 'Закрыт' загружен из .env: {CLOSED_STATUS_IDS}")
        return CLOSED_STATUS_IDS

    logger.info("🔍 Определение ID статуса 'Закрыт' в Jira...")

    try:
        jira = get_jira_connection()
        statuses = jira.statuses()

        closed_ids = []
        for status in statuses:
            if status.name.lower() in ['закрыт', 'closed', 'закрыто']:
                closed_ids.append(status.id)
                logger.info(f"   📌 Найден статус: {status.name} (ID: {status.id})")

        if closed_ids:
            save_closed_status_ids(closed_ids)
            logger.info(f"✅ ID сохранены в .env: {closed_ids}")
            return closed_ids
        else:
            logger.warning("⚠️  Статус 'Закрыт' не найден.")
            return []

    except Exception as e:
        logger.error(f"❌ Ошибка определения статуса: {e}")
        return []


def save_closed_status_ids(status_ids: List[str]) -> None:
    """
    Сохраняет ID статусов в файл .env.

    Args:
        status_ids: Список ID для сохранения
    """
    # .env находится в корневой директории (на уровень выше core/)
    env_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '.env')

    env_content = ''
    detected_encoding = 'utf-8'
    
    if os.path.exists(env_path):
        # Читаем в UTF-8, если не получается — пробуем cp1251 (Windows)
        for encoding in ['utf-8', 'cp1251']:
            try:
                with open(env_path, 'r', encoding=encoding) as f:
                    env_content = f.read()
                detected_encoding = encoding
                break
            except UnicodeDecodeError:
                continue

    if 'CLOSED_STATUS_IDS=' in env_content:
        # Заменяем только значение после CLOSED_STATUS_IDS=
        lines = env_content.split('\n')
        new_lines = []
        for line in lines:
            if line.startswith('CLOSED_STATUS_IDS='):
                new_lines.append(f'CLOSED_STATUS_IDS={",".join(status_ids)}')
            else:
                new_lines.append(line)
        env_content = '\n'.join(new_lines)
    else:
        # Добавляем в конец файла
        env_content = env_content.rstrip() + f'\nCLOSED_STATUS_IDS={",".join(status_ids)}\n'

    # Пишем в UTF-8 без BOM (стандарт для python-dotenv)
    with open(env_path, 'w', encoding='utf-8', newline='\n') as f:
        f.write(env_content)

# Импортируем функции проверки проблем из справочника
from core.problems_dict import (
    check_no_assignee,
    check_no_time_spent,
    check_no_resolution_date,
    check_incorrect_status,
    check_overdue,
    check_late_creation,
    check_inactive,
    PROBLEM_TYPES,
)

def validate_issue(issue: Any, jira: Optional[JIRA] = None, closed_status_ids: Optional[List[str]] = None, project_key: Optional[str] = None) -> List[str]:
    """
    Проверяет задачу на корректность заполнения.

    Использует справочник проблем (core.problems_dict) для проверки.

    Args:
        issue: Объект задачи Jira
        jira: Объект подключения к Jira (нужен для проверки changelog)
        closed_status_ids: Список ID закрытых статусов (опционально)
        project_key: Ключ проекта (опционально, для исключений)

    Returns:
        List[str]: Список проблем
    """
    problems = []

    # Используем переданный список или глобальный
    status_ids = closed_status_ids if closed_status_ids else CLOSED_STATUS_IDS

    # Проверка: нет исполнителя
    if check_no_assignee(issue):
        problems.append(PROBLEM_TYPES['NO_ASSIGNEE']['short_name'])

    # Проверка: нет фактического времени (исключение для проектов из EXCLUDED_PROJECTS_NO_TIMESPENT)
    if check_no_time_spent(issue):
        # Проверяем, не в исключённом ли проекте задача
        if not project_key or project_key.upper() not in [p.upper() for p in EXCLUDED_PROJECTS_NO_TIMESPENT]:
            problems.append(PROBLEM_TYPES['NO_TIME_SPENT']['short_name'])

    # Проверка: нет даты решения
    if check_no_resolution_date(issue):
        problems.append(PROBLEM_TYPES['NO_RESOLUTION_DATE']['short_name'])

    # Проверка: просрочка планирования (создана позже даты решения)
    if issue.fields.created and issue.fields.duedate:
        threshold = PROBLEM_TYPES['LATE_CREATION'].get('threshold_days', 7)
        if check_late_creation(issue, threshold):
            try:
                created_date = datetime.strptime(issue.fields.created[:10], '%Y-%m-%d')
                due_date = datetime.strptime(issue.fields.duedate[:10], '%Y-%m-%d')
                days_diff = (created_date - due_date).days
                problems.append(f"Создана на {days_diff} дн. позже даты решения")
            except Exception:
                pass  # Если не удалось сравнить даты — не считаем проблемой

    # Проверка: просрочена (дата решения истёк)
    if check_overdue(issue):
        problems.append(PROBLEM_TYPES['OVERDUE']['short_name'])

    # Проверка: не двигается (неактивна)
    threshold_inactive = PROBLEM_TYPES['INACTIVE'].get('threshold_days', RISK_ZONE_INACTIVITY_THRESHOLD)
    if check_inactive(issue, threshold_inactive):
        try:
            updated = datetime.strptime(issue.fields.updated[:19], '%Y-%m-%dT%H:%M:%S')
            days_inactive = (datetime.now() - updated).days
            problems.append(f"Не двигается {days_inactive} дн.")
        except Exception:
            pass

    # Проверка статуса "Закрыт" по ID
    if issue.fields.status:
        status_id = issue.fields.status.id
        status_name = issue.fields.status.name

        # Проверяем changelog ТОЛЬКО если статус "Закрыт"
        if status_id in status_ids:
            is_correct_close = False

            # Проверяем, не является ли исполнитель исключением (holin и т.п.)
            assignee_name = ''
            if issue.fields.assignee:
                assignee_name = issue.fields.assignee.name if hasattr(issue.fields.assignee, 'name') else issue.fields.assignee.displayName

            for exc in EXCLUDED_ASSIGNEE_CLOSE:
                if exc.lower() in assignee_name.lower():
                    is_correct_close = True
                    break

            # Если не исключение, проверяем changelog (кто перевёл в "Закрыт")
            # Используем предзагруженный changelog из issues_normal (экономия запросов к API)
            if not is_correct_close:
                try:
                    # Проверяем, есть ли changelog в предзагруженном объекте
                    if hasattr(issue, 'changelog') and issue.changelog:
                        # Ищем последний переход в статус "Закрыт"
                        found_correct_close = False
                        for history in reversed(issue.changelog.histories):
                            for item in history.items:
                                if item.field == 'status' and item.toString:
                                    # Проверяем, был ли это переход в закрытый статус
                                    if hasattr(item, 'to') and item.to in CLOSED_STATUS_IDS:
                                        # Проверяем, кто сделал переход
                                        author_name = ''
                                        if hasattr(history, 'author') and history.author:
                                            author_name = history.author.name if hasattr(history.author, 'name') else history.author.displayName

                                        # Если переход сделал пользователь демона — это корректно
                                        if JIRA_USER and JIRA_USER.lower() in author_name.lower():
                                            is_correct_close = True
                                            found_correct_close = True
                                        break
                            if found_correct_close:
                                break
                    else:
                        # Changelog отсутствует — это проблема
                        logger.warning(f"⚠️  Отсутствует changelog для {issue.key}")
                        problems.append('Не удалось проверить историю переходов')
                except Exception as e:
                    # Если не удалось получить changelog, считаем это проблемой
                    logger.warning(f"⚠️  Ошибка при проверке changelog для {issue.key}: {e}")
                    problems.append('Не удалось проверить историю переходов')

            # Если статус "Закрыт" и не корректно закрыт — это проблема
            if not is_correct_close:
                problems.append(PROBLEM_TYPES['INCORRECT_STATUS']['short_name'])

    return problems


def get_column_order(block: str, extra_verbose: bool = False) -> List[str]:
    """
    Возвращает порядок колонок для каждого блока.

    Args:
        block: Название блока отчёта
        extra_verbose: Показывать ли ID объектов

    Returns:
        List[str]: Список названий колонок

    Raises:
        ValueError: Если блок не найден
    """
    if block == 'summary':
        if extra_verbose:
            return ['Клиент (Проект) [project]', 'ID [id]', 'Задач закрыто', 'Корректных', 'С ошибками', 'Оценка (ч) [timeoriginalestimate]', 'Факт (ч) [timespent]', 'Отклонение']
        return ['Клиент (Проект)', 'Задач закрыто', 'Корректных', 'С ошибками', 'Оценка (ч)', 'Факт (ч)', 'Отклонение']
    elif block == 'assignees':
        if extra_verbose:
            return ['Исполнитель [assignee]', 'ID [accountId]', 'Задач', 'Корректных', 'С ошибками', 'Оценка (ч) [timeoriginalestimate]', 'Факт (ч) [timespent]', 'Отклонение']
        return ['Исполнитель', 'Задач', 'Корректных', 'С ошибками', 'Оценка (ч)', 'Факт (ч)', 'Отклонение']
    elif block == 'detail':
        if extra_verbose:
            return ['URL', 'ID [id]', 'Дата решения [resolutiondate]', 'Дата исполнения [duedate]', 'Дата создания [created]', 'Проект [project]', 'Статус [status]', 'Задача [summary]', 'Исполнитель [assignee]', 'Факт (ч) [timespent]', 'Тип [issuetype]']
        return ['URL', 'Дата решения', 'Дата исполнения', 'Дата создания', 'Проект', 'Статус', 'Задача', 'Исполнитель', 'Факт (ч)', 'Тип']
    elif block == 'issues':
        if extra_verbose:
            return ['URL', 'ID [id]', 'Дата исполнения [duedate]', 'Дата создания [created]', 'Проект [project]', 'Задача [summary]', 'Исполнитель [assignee]', 'Автор [creator]', 'Проблемы']
        return ['URL', 'Дата исполнения', 'Дата создания', 'Проект', 'Задача', 'Исполнитель', 'Автор', 'Проблемы']
    elif block == 'internal':
        if extra_verbose:
            return ['URL', 'ID [id]', 'Проект ID [project.id]', 'Проект [project]', 'Ключ [key]', 'Задача [summary]', 'Исполнитель [assignee]', 'Статус [status]', 'Факт (ч) [timespent]', 'Дата создания [created]', 'Дата исполнения [duedate]', 'Тип [issuetype]']
        return ['URL', 'Проект', 'Ключ', 'Задача', 'Исполнитель', 'Статус', 'Факт (ч)', 'Дата создания', 'Дата исполнения', 'Тип']
    elif block == 'risk_zone':
        if extra_verbose:
            return ['URL', 'Ключ [key]', 'Задача [summary]', 'Исполнитель [assignee]', 'Статус [status]', 'Факторы риска', 'Приоритет [priority]']
        return ['URL', 'Ключ', 'Задача', 'Исполнитель', 'Статус', 'Факторы риска', 'Приоритет']
    else:
        logger.warning(f"⚠️  Неизвестный блок '{block}', используются колонки по умолчанию")
        return ['Проект', 'Ключ', 'Задача', 'Исполнитель', 'Статус', 'Дата создания', 'Дата исполнения', 'Факт (ч)', 'Оценка (ч)']
        
def generate_report(
    project_keys: Optional[Union[str, List[str]]] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    days: int = 30,
    assignee_filter: Optional[Union[str, List[str]]] = None,
    issue_types: Optional[Union[str, List[str]]] = None,
    blocks: Optional[List[str]] = None,
    verbose: bool = False,
    extra_verbose: bool = False,
    closed_status_ids: Optional[List[str]] = None,
    include_risk_zone: bool = True  # Новый параметр для Risk Zone
) -> Dict[str, Any]:
    """
    Генерирует отчёт по задачам Jira.

    Args:
        project_keys: Ключ проекта или список проектов (None = все проекты)
        start_date: Дата начала в формате ГГГГ-ММ-ДД (None = прошлый месяц)
        end_date: Дата окончания в формате ГГГГ-ММ-ДД (имеет приоритет над days)
        days: Количество дней для отчёта (0 = без ограничений)
        assignee_filter: Фильтр по исполнителю или список исполнителей
        issue_types: Фильтр по типам задач или список типов
        blocks: Список блоков отчёта (None = все)
        verbose: Режим отладки
        extra_verbose: Показывать ID объектов

    Returns:
        Dict[str, Any]: Словарь с данными отчёта
    """
    # Авто-определение ID статуса "Закрыт" (локальная переменная, не мутим глобальную)
    closed_status_ids = CLOSED_STATUS_IDS
    if not closed_status_ids or closed_status_ids[0] == '':
        closed_status_ids = get_closed_status_ids()

    # Нормализация множественных фильтров
    project_keys = normalize_filter(project_keys, upper=True)
    assignee_filter = normalize_filter(assignee_filter)
    issue_types = normalize_filter(issue_types)

    # Обработка дат: end_date имеет приоритет над days
    if start_date:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
    else:
        start_date_obj = get_default_start_date()

    start_date_str = start_date_obj.strftime('%Y-%m-%d')

    # end_date имеет приоритет над days
    if end_date:
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
        end_date_str = end_date_obj.strftime('%Y-%m-%d')
        # Для проблемных задач: +2 месяца к end_date
        issues_end_obj = end_date_obj + relativedelta(months=2)
        issues_end_str = issues_end_obj.strftime('%Y-%m-%d')
    elif days > 0:
        end_date_obj = start_date_obj + timedelta(days=days - 1)
        end_date_str = end_date_obj.strftime('%Y-%m-%d')
        # Для проблемных задач: +2 месяца к концу периода
        issues_end_obj = start_date_obj + timedelta(days=days) + relativedelta(months=2)
        issues_end_str = issues_end_obj.strftime('%Y-%m-%d')
    else:
        # Без ограничений по датам - используем текущую дату как конец
        end_date_obj = datetime.now()
        end_date_str = end_date_obj.strftime('%Y-%m-%d')
        issues_end_str = end_date_str

    jira = get_jira_connection()

    # Список проектов
    if project_keys and len(project_keys) > 0:
        # Фильтр по выбранным проектам
        projects_map = {}
        for proj_key in project_keys:
            try:
                # Пробуем использовать кэширующую функцию из web/app.py
                # Если она доступна (веб-режим), иначе используем прямой запрос
                try:
                    from web.app import get_project_cached
                    proj = get_project_cached(jira, proj_key)
                except ImportError:
                    # Консольный режим - кэш не доступен
                    proj = jira.project(proj_key)
                
                if proj:
                    projects_map[proj.key] = proj.name
            except Exception:
                logger.warning(f"Проект {proj_key} не найден")
        projects_keys = list(projects_map.keys())  # Обновляем список ключей
    else:
        # Все проекты
        all_projects = jira.projects()
        projects_map = {}
        for proj in all_projects:
            if proj.key in EXCLUDED_PROJECTS:
                continue
            if hasattr(proj, 'archived') and proj.archived:
                continue
            projects_map[proj.key] = proj.name
        projects_keys = list(projects_map.keys())

    all_issues_data = []
    summary_data = []
    issues_with_problems = []
    all_issues_normal = []  # Список всех задач для Risk Zone

    # Формируем фильтр по типам задач для JQL
    issue_type_filter = ''
    if issue_types and len(issue_types) > 0:
        # Санизируем каждый тип задачи (только разрешённые символы)
        sanitized_types = []
        for t in issue_types:
            try:
                sanitized_types.append(sanitize_jql_identifier(t))
            except ValueError as e:
                logging.warning(f"Пропущен недопустимый тип задачи '{t}': {e}")
        if sanitized_types:
            issue_type_filter = ' AND issuetype IN (' + ','.join(sanitized_types) + ')'

    # Формируем фильтр по исполнителям для JQL
    assignee_filter_jql = ''
    if assignee_filter and len(assignee_filter) > 0:
        # Санизируем каждый username (только разрешённые символы)
        sanitized_assignees = []
        for a in assignee_filter:
            try:
                sanitized_assignees.append(sanitize_jql_identifier(a))
            except ValueError as e:
                logging.warning(f"Пропущен недопустимый пользователь '{a}': {e}")
        if sanitized_assignees:
            assignee_filter_jql = ' AND assignee IN (' + ','.join(sanitized_assignees) + ')'

    # ========== ОПТИМИЗАЦИЯ: 2 глобальных запроса вместо 2×N проектов ==========
    # Санизируем даты (только формат YYYY-MM-DD)
    start_date_safe = sanitize_jql_string_literal(start_date_str)
    end_date_safe = sanitize_jql_string_literal(end_date_str)
    issues_end_safe = sanitize_jql_string_literal(issues_end_str)

    # Формируем список проектов для JQL
    projects_jql = ','.join([sanitize_jql_identifier(p) for p in projects_keys])
    project_filter = f"project IN ({projects_jql})" if projects_keys else ""

    # Глобальный JQL для обычных отчётов (фильтр по duedate)
    if days > 0:
        jql_normal_global = (f"{project_filter} "
                          f"AND duedate >= '{start_date_safe}' "
                          f"AND duedate <= '{end_date_safe}' "
                          f"AND duedate is not null"
                          f"{issue_type_filter}"
                          f"{assignee_filter_jql} "
                          f"ORDER BY duedate ASC")
    else:
        jql_normal_global = (f"{project_filter} "
                          f"AND duedate is not null"
                          f"{issue_type_filter}"
                          f"{assignee_filter_jql} "
                          f"ORDER BY duedate DESC")

    # Глобальный JQL для проблемных задач (фильтр по created + 2 месяца)
    if days > 0:
        jql_issues_global = (f"{project_filter} "
                          f"AND created >= '{start_date_safe}' "
                          f"AND created <= '{issues_end_safe}'"
                          f"{issue_type_filter}"
                          f"{assignee_filter_jql} "
                          f"ORDER BY created ASC")
    else:
        jql_issues_global = (f"{project_filter} "
                          f"AND created is not null"
                          f"{issue_type_filter}"
                          f"{assignee_filter_jql} "
                          f"ORDER BY created DESC")

    # ========== ПОЛНАЯ ОПТИМИЗАЦИЯ: 2 REST API запроса + 1 проход ==========

    logger.info(f"🚀 Оптимизация: выполнение 2 REST API запросов вместо {len(projects_keys) * 2}")

    # Получаем все задачи через REST API (с creator!)
    issues_all_global = fetch_issues_via_rest(jira, jql_issues_global)
    issues_normal_global = fetch_issues_via_rest(jira, jql_normal_global)

    # Добавляем все задачи normal в список для Risk Zone
    all_issues_normal.extend(issues_normal_global)

    # ========== Обработка ВСЕХ задач за ОДИН проход ==========
    # Словарь для агрегации по проектам: proj_key -> {spent, estimated, correct, issues}
    project_stats = {}

    for issue_data in issues_normal_global:
        # REST API возвращает dict, а не объект
        fields = issue_data.get('fields', {})
        proj_key = fields.get('project', {}).get('key', '')
        proj_name = projects_map.get(proj_key, proj_key)

        # Получаем значения полей из REST API ответа
        timespent = fields.get('timespent')
        timeoriginalestimate = fields.get('timeoriginalestimate')
        spent = convert_seconds_to_hours(timespent)
        estimated = convert_seconds_to_hours(timeoriginalestimate)

        issuetype = fields.get('issuetype', {})
        issue_type = issuetype.get('name', 'Задача') if issuetype else 'Задача'

        assignee = fields.get('assignee', {}).get('displayName', 'Без исполнителя') if fields.get('assignee') else 'Без исполнителя'

        duedate = fields.get('duedate', '-') or '-'
        if duedate and duedate != '-':
            duedate = duedate[:10]

        resolutiondate = fields.get('resolutiondate', '-') or '-'
        if resolutiondate and resolutiondate != '-':
            resolutiondate = resolutiondate[:10]

        created = fields.get('created', '-') or '-'
        if created and created != '-':
            created = created[:10]

        status = fields.get('status', {})
        status_name = status.get('name', '-') if status else '-'
        status_category = status.get('statusCategory', {}).get('key', '-') if status else '-'
        status_full = f"{status_name} ({status_category})"

        issue_key = issue_data.get('key', '')
        issue_url = f"{JIRA_SERVER}/browse/{issue_key}"
        issue_id = issue_data.get('id') if extra_verbose else None

        # Создаём псевдо-объект issue для validate_issue
        class MockIssue:
            def __init__(self, data):
                self.fields = type('obj', (object,), {
                    'assignee': type('obj', (object,), {'displayName': data.get('fields', {}).get('assignee', {}).get('displayName') if data.get('fields', {}).get('assignee') else None})(),
                    'timespent': data.get('fields', {}).get('timespent'),
                    'timeoriginalestimate': data.get('fields', {}).get('timeoriginalestimate'),
                    'resolutiondate': data.get('fields', {}).get('resolutiondate'),
                    'status': type('obj', (object,), {
                        'id': data.get('fields', {}).get('status', {}).get('id'),
                        'name': data.get('fields', {}).get('status', {}).get('name'),
                        'statusCategory': type('obj', (object,), {'key': data.get('fields', {}).get('status', {}).get('statusCategory', {}).get('key')})()
                    })(),
                    'created': data.get('fields', {}).get('created'),
                    'duedate': data.get('fields', {}).get('duedate'),
                    'issuetype': type('obj', (object,), {'name': data.get('fields', {}).get('issuetype', {}).get('name')})()
                })()
                self.key = data.get('key', '')

        mock_issue = MockIssue(issue_data)
        problems = validate_issue(mock_issue, jira, closed_status_ids, proj_key)

        # Формируем отображаемые значения с ID если нужно
        project_display = proj_name
        status_display = f"{status_name} ({status_category}) [{status_id}]"  # Всегда показываем ID статуса
        issue_type_display = issue_type
        assignee_display = assignee

        if extra_verbose:
            project_id = fields.get('project', {}).get('id', '')
            project_display = f"{proj_name} [{project_id}]" if project_id else proj_name

            type_id = issuetype.get('id', '')
            issue_type_display = f"{issue_type} [{type_id}]" if type_id else issue_type

            assignee_id = fields.get('assignee', {}).get('accountId', '')
            assignee_display = f"{assignee} [{assignee_id}]" if assignee_id else assignee

        issue_data = {
            'URL': issue_url,
            'ID': issue_id,
            'Проект': project_display,
            'Ключ': issue_key,
            'Тип': issue_type_display,
            'Задача': fields.get('summary', ''),
            'Исполнитель': assignee_display,
            'Статус': status_display,
            'Дата создания': created,
            'Дата исполнения': duedate,
            'Дата решения': resolutiondate,
            'Факт (ч)': spent,
            'Оценка (ч)': estimated,
            'Проблемы': ', '.join(problems) if problems else ''
        }

        all_issues_data.append(issue_data)

        # Агрегация статистики по проекту
        if proj_key not in project_stats:
            project_stats[proj_key] = {'name': proj_name, 'spent': 0.0, 'estimated': 0.0, 'correct': 0, 'issues': 0}

        if not problems:
            project_stats[proj_key]['spent'] += spent
            project_stats[proj_key]['estimated'] += estimated
            project_stats[proj_key]['correct'] += 1
        else:
            project_stats[proj_key]['issues'] += 1

        # Проблемные задачи — берём creator из REST API
        if problems:
            creator = fields.get('creator', {})
            author = creator.get('displayName', 'N/A') if creator else 'N/A'
            author_id = creator.get('accountId', '') if creator else ''

            if not creator:
                logger.warning(f"⚠️  Creator не доступен для {issue_key}")

            author_display = f"{author} [{author_id}]" if extra_verbose and author_id else author

            issue_data_probs = {
                'URL': issue_url,
                'URL_debug': f"/?debug={issue_key}",
                'Проект': proj_name,
                'Задача': fields.get('summary', ''),
                'Исполнитель': assignee,
                'Автор': author_display,
                'Дата создания': created,
                'Дата исполнения': duedate,
                'Проблемы': ', '.join(problems)
            }
            if extra_verbose:
                issue_data_probs = {'ID': issue_data.get('id'), **issue_data_probs}
            issues_with_problems.append(issue_data_probs)

    # Формируем summary из агрегированной статистики
    for proj_key, stats in project_stats.items():
        if stats['correct'] > 0 or stats['issues'] > 0:
            if extra_verbose:
                # Берём ID проекта из первой задачи
                proj_issues_list = [i for i in issues_normal_global if i.get('fields', {}).get('project', {}).get('key') == proj_key]
                proj_id = proj_issues_list[0].get('fields', {}).get('project', {}).get('id', '') if proj_issues_list else ''
                summary_row = {
                    'Клиент (Проект)': stats['name'],
                    'ID': proj_id,
                    'Задач закрыто': stats['correct'] + stats['issues'],
                    'Корректных': stats['correct'],
                    'С ошибками': stats['issues'],
                    'Оценка (ч)': round(stats['estimated'], 2),
                    'Факт (ч)': round(stats['spent'], 2),
                    'Отклонение': round(stats['estimated'] - stats['spent'], 2)
                }
            else:
                summary_row = {
                    'Клиент (Проект)': stats['name'],
                    'Задач закрыто': stats['correct'] + stats['issues'],
                    'Корректных': stats['correct'],
                    'С ошибками': stats['issues'],
                    'Оценка (ч)': round(stats['estimated'], 2),
                    'Факт (ч)': round(stats['spent'], 2),
                    'Отклонение': round(stats['estimated'] - stats['spent'], 2)
                }
            summary_data.append(summary_row)
    
    df_detail = pd.DataFrame(all_issues_data)
    df_summary = pd.DataFrame(summary_data)
    df_issues = pd.DataFrame(issues_with_problems)
    
    # Сортировка и группировка
    if not df_detail.empty:
        df_detail = df_detail.sort_values(by=['Тип', 'Проект', 'Дата решения'], ascending=[True, True, True])

        # Группировка по исполнителям - ИСКЛЮЧАЕМ "Без исполнителя"
        if not df_detail.empty:
            # Фильтруем только задачи с исполнителем (не "Без исполнителя")
            df_with_assignee = df_detail[~df_detail['Исполнитель'].str.contains('Без исполнителя', na=False)]
            
            if not df_with_assignee.empty:
                df_assignees = df_with_assignee.groupby('Исполнитель').agg(
                    tasks_count=('Ключ', 'count'),
                    correct_count=('Проблемы', lambda x: (x == '').sum()),
                    issues_count=('Проблемы', lambda x: (x != '').sum()),
                    fact_sum=('Факт (ч)', 'sum'),
                    estimate_sum=('Оценка (ч)', 'sum')
                ).reset_index()
                # Переименовываем колонки обратно в кириллицу для отображения
                df_assignees = df_assignees.rename(columns={
                    'tasks_count': 'Задач',
                    'correct_count': 'Корректных',
                    'issues_count': 'С ошибками',
                    'fact_sum': 'Факт (ч)',
                    'estimate_sum': 'Оценка (ч)'
                })
                df_assignees['Отклонение'] = df_assignees['Оценка (ч)'] - df_assignees['Факт (ч)']
                df_assignees = df_assignees.round(2)
                df_assignees = df_assignees.sort_values(by='Факт (ч)', ascending=False)

                # Добавляем колонку ID для extra_verbose (извлекаем из "Исполнитель [ID]")
                if extra_verbose:
                    def extract_id(name):
                        if '[' in name and ']' in name:
                            return name.split('[')[-1].split(']')[0]
                        return ''
                    df_assignees.insert(1, 'ID', df_assignees['Исполнитель'].apply(extract_id))
            else:
                df_assignees = pd.DataFrame()
        else:
            df_assignees = pd.DataFrame()
    else:
        df_assignees = pd.DataFrame()

    # Блок "Непонятное" - задачи из внутренних проектов (NEW, local)
    df_internal = pd.DataFrame()
    if 'internal' in (blocks or list(REPORT_BLOCKS.keys())) and INTERNAL_PROJECTS:
        internal_issues_data = []
        for internal_proj_key in INTERNAL_PROJECTS:
            if internal_proj_key not in projects_map:
                continue
            internal_proj_name = projects_map[internal_proj_key]
            
            # Получаем все задачи из внутреннего проекта за период
            jql_internal = (f"project = {internal_proj_key} "
                          f"AND created >= '{start_date_str}' "
                          f"AND created <= '{end_date_str}' "
                          f"ORDER BY created ASC")
            
            try:
                internal_issues = jira.search_issues(jql_internal, maxResults=False,
                                                     fields='summary, assignee, timespent, timeoriginalestimate, resolutiondate, issuetype, duedate, status, created')
                
                for issue in internal_issues:
                    spent = convert_seconds_to_hours(issue.fields.timespent)
                    estimated = convert_seconds_to_hours(issue.fields.timeoriginalestimate)
                    
                    issue_type = issue.fields.issuetype.name if issue.fields.issuetype else 'Задача'
                    assignee = issue.fields.assignee.displayName if issue.fields.assignee else 'Без исполнителя'
                    status_name = issue.fields.status.name if issue.fields.status else '-'
                    created = issue.fields.created[:10] if issue.fields.created else '-'
                    
                    issue_url = f"{JIRA_SERVER}/browse/{issue.key}"
                    
                    internal_issues_data.append({
                        'URL': issue_url,
                        'Проект': internal_proj_name,
                        'Ключ': issue.key,
                        'Тип': issue_type,
                        'Задача': issue.fields.summary,
                        'Исполнитель': assignee,
                        'Статус': status_name,
                        'Дата создания': created,
                        'Факт (ч)': spent,
                        'Оценка (ч)': estimated
                    })
            except Exception as e:
                logger.warning(f"Не удалось получить задачи из проекта {internal_proj_key}: {e}")
        
        if internal_issues_data:
            df_internal = pd.DataFrame(internal_issues_data)
            df_internal = df_internal.sort_values(by='Дата создания', ascending=True)

    result = {
        'period': f"{start_date_str} — {end_date_str}",
        'blocks': blocks or list(REPORT_BLOCKS.keys()),
        'total_projects': len(df_summary),
        'total_tasks': len(df_detail),
        'total_correct': len(df_detail[df_detail['Проблемы'] == '']) if not df_detail.empty else 0,
        'total_issues': len(df_issues),
        'total_spent': df_summary['Факт (ч)'].sum() if not df_summary.empty else 0,
        'total_estimated': df_summary['Оценка (ч)'].sum() if not df_summary.empty else 0,
    }
    
    if 'summary' in result['blocks']:
        result['summary'] = df_summary
    if 'assignees' in result['blocks']:
        result['assignees'] = df_assignees
    if 'detail' in result['blocks']:
        result['detail'] = df_detail
    if 'issues' in result['blocks']:
        result['issues'] = df_issues
    if 'internal' in result['blocks']:
        result['internal'] = df_internal

    # Фильтрация колонок для каждого блока
    if 'summary' in result['blocks'] and not result['summary'].empty:
        cols = get_column_order('summary', extra_verbose)
        available_cols = [c for c in cols if c in result['summary'].columns]
        result['summary'] = result['summary'][available_cols]

    if 'assignees' in result['blocks'] and not result['assignees'].empty:
        cols = get_column_order('assignees', extra_verbose)
        available_cols = [c for c in cols if c in result['assignees'].columns]
        result['assignees'] = result['assignees'][available_cols]

    if 'detail' in result['blocks'] and not result['detail'].empty:
        cols = get_column_order('detail', extra_verbose)
        available_cols = [c for c in cols if c in result['detail'].columns]
        result['detail'] = result['detail'][available_cols]

    if 'issues' in result['blocks'] and not result['issues'].empty:
        cols = get_column_order('issues', extra_verbose)
        available_cols = [c for c in cols if c in result['issues'].columns]
        result['issues'] = result['issues'][available_cols]

    if 'internal' in result['blocks'] and not result['internal'].empty:
        cols = get_column_order('internal', extra_verbose)
        available_cols = [c for c in cols if c in result['internal'].columns]
        result['internal'] = result['internal'][available_cols]

    # ========== БЛОК "RISK ZONE" - ЗАВИСШИЕ ЗАДАЧИ ==========
    if include_risk_zone:
        risk_issues = []
        today = datetime.now()

        logger.info(f"🔍 Проверка Risk Zone... all_issues_normal: {len(all_issues_normal)} задач")

        # Проверяем все задачи из всех проектов
        if all_issues_normal:
            logger.info(f"   Найдено {len(all_issues_normal)} задач для проверки")
            for issue_data in all_issues_normal:
                # REST API возвращает dict, а не объект
                fields = issue_data.get('fields', {})
                issue_key = issue_data.get('key', '')

                risk_factors = []

                # Получаем статус для проверки
                status = fields.get('status', {})
                status_id = status.get('id', '')

                # 1. Задачи без исполнителя
                assignee = fields.get('assignee')
                if not assignee:
                    risk_factors.append('Без исполнителя')

                # 2. Задачи с истёкшим сроком (Due Date)
                # Проверка: duedate < сегодня И status.id НЕ в закрытых статусах
                duedate = fields.get('duedate')
                if duedate:
                    due_date = datetime.strptime(duedate[:10], '%Y-%m-%d')
                    if due_date < today and status_id not in CLOSED_STATUS_IDS:
                        days_overdue = (today - due_date).days
                        risk_factors.append(f'Просрочена на {days_overdue} дн.')

                # 3. Задачи, которые не двигались > порога неактивности
                # Проверка: не обновлялась N дней И status.id НЕ в закрытых статусах
                updated = fields.get('updated')
                if updated:
                    updated_dt = datetime.strptime(updated[:19], '%Y-%m-%dT%H:%M:%S')
                    days_inactive = (today - updated_dt).days
                    if days_inactive > RISK_ZONE_INACTIVITY_THRESHOLD and status_id not in CLOSED_STATUS_IDS:
                        risk_factors.append(f'Не двигается {days_inactive} дн.')

                # Если есть факторы риска - добавляем в отчёт
                if risk_factors:
                    assignee_name = assignee.get('displayName', 'Без исполнителя') if assignee else 'Без исполнителя'
                    priority = fields.get('priority', {})
                    risk_issues.append({
                        'URL': f"{JIRA_SERVER}/browse/{issue_key}",
                        'Ключ': issue_key,
                        'Задача': fields.get('summary', ''),
                        'Исполнитель': assignee_name,
                        'Статус': status_name,
                        'Факторы риска': '; '.join(risk_factors),
                        'Приоритет': priority.get('name', 'Normal')
                    })

            logger.info(f"   Найдено {len(risk_issues)} рисковых задач")
        else:
            logger.warning("   ⚠️  all_issues_normal пуст")

        if risk_issues:
            result['risk_zone'] = pd.DataFrame(risk_issues)
            result['risk_zone'] = result['risk_zone'].sort_values('Приоритет', ascending=False)
            logger.info(f"✅ Risk Zone заполнен: {len(risk_issues)} задач")
        else:
            logger.info("ℹ️  Risk Zone пуст (нет рисковых задач)")

    return result

def generate_excel(report_data: Dict[str, Any], output: Optional[Union[str, io.BytesIO]] = None) -> Union[str, io.BytesIO]:
    """
    Выгружает отчёт в Excel с форматированием.

    Args:
        report_data: Данные отчёта
        output: Путь к файлу или BytesIO объект

    Returns:
        Union[str, io.BytesIO]: Имя файла или BytesIO объект
        
    Raises:
        ValueError: Если отчёт превышает MAX_EXCEL_ROWS строк
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Проверка на максимальный размер
    total_rows = 0
    for key in ['summary', 'assignees', 'detail', 'issues', 'internal', 'risk_zone']:
        df = report_data.get(key)
        if df is not None and not df.empty:
            total_rows += len(df)
    
    if total_rows > MAX_EXCEL_ROWS:
        raise ValueError(
            f'Отчёт слишком большой: {total_rows} строк (максимум {MAX_EXCEL_ROWS}). '
            'Разбейте отчёт на несколько периодов.'
        )

    if output is None:
        output = f"jira_report_{report_data['period'].replace(' — ', '_to_').replace(' ', '')}.xlsx"

    writer = pd.ExcelWriter(output, engine='openpyxl')

    try:
        # Словарь с данными для каждого листа
        sheets_data = {
            'Сводка': report_data.get('summary'),
            'Исполнители': report_data.get('assignees'),
            'Детали': report_data.get('detail'),
            'Проблемы': report_data.get('issues'),
            'Непонятное': report_data.get('internal'),
            'Risk Zone': report_data.get('risk_zone')
        }
        
        for sheet_name, df in sheets_data.items():
            if df is not None and not df.empty:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Форматирование
                worksheet = writer.sheets[sheet_name]
                
                # Стили
                header_font = Font(bold=True, color='FFFFFF', size=11)
                header_fill = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                
                # Форматирование заголовков
                for col in worksheet[1]:
                    col.font = header_font
                    col.fill = header_fill
                    col.alignment = header_alignment
                
                # Автоширина колонок
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column].width = adjusted_width
                
                # Выделение первой строки (заголовки)
                worksheet.freeze_panes = 'A2'
                
    finally:
        writer.close()

    return output

# =============================================
# КОНСОЛЬНЫЙ ЗАПУСК
# =============================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description='Генерация отчёта по закрытым задачам из Jira',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
БЛОКИ ОТЧЁТА:
  summary   - Сводка по проектам
  assignees - Нагрузка по исполнителям
  detail    - Детализация по задачам
  issues    - Проблемные задачи

ПРИМЕРЫ:
  python3 jira_report.py -e
  python3 jira_report.py -b summary,assignees -e
  python3 jira_report.py -p WEB -a "Иванов" -b detail -e
  python3 jira_report.py -b issues -vv
        '''
    )
    parser.add_argument('-p', '--project', type=str, help='Ключ проекта')
    parser.add_argument('-s', '--start-date', type=str, help='Дата начала (ГГГГ-ММ-ДД)')
    parser.add_argument('-d', '--days', type=int, default=30, help='Период в днях')
    parser.add_argument('-a', '--assignee', type=str, help='Фильтр по исполнителю')
    parser.add_argument('-b', '--blocks', type=str, help='Блоки отчёта (через запятую)')
    parser.add_argument('-e', '--excel', action='store_true', help='Выгрузка в Excel')
    parser.add_argument('-v', '--verbose', action='store_true', help='Режим отладки')
    parser.add_argument('-vv', '--extra-verbose', action='store_true', help='Показывать ID задач во всех отчётах')
    args = parser.parse_args()
    
    blocks = None
    if args.blocks:
        blocks = [b.strip() for b in args.blocks.split(',')]
        invalid = [b for b in blocks if b not in REPORT_BLOCKS]
        if invalid:
            print(f"❌ Неверные блоки: {invalid}")
            print(f"Доступные: {list(REPORT_BLOCKS.keys())}")
            sys.exit(1)
    
    # Авто-определение статуса перед запуском (локальная переменная)
    closed_status_ids = CLOSED_STATUS_IDS
    if not closed_status_ids or closed_status_ids[0] == '':
        closed_status_ids = get_closed_status_ids()

    print(f"🔌 Генерация отчёта...")
    if args.blocks:
        print(f"📦 Блоки: {', '.join(blocks)}")

    report = generate_report(
        project_keys=args.project,
        start_date=args.start_date,
        days=args.days,
        assignee_filter=args.assignee,
        blocks=blocks,
        verbose=args.verbose,
        extra_verbose=args.extra_verbose,
        closed_status_ids=closed_status_ids
    )
    
    print("\n" + "="*100)
    print(f"📋 ОТЧЁТ ЗА {report['period']}")
    print("="*100)
    
    if 'summary' in report:
        print("\n📊 СВОДКА ПО ПРОЕКТАМ:")
        print("="*100)
        print(report['summary'].to_string(index=False))
    
    if 'assignees' in report and not report['assignees'].empty:
        print("\n👤 НАГРУЗКА ПО ИСПОЛНИТЕЛЯМ:")
        print("="*100)
        print(report['assignees'].to_string(index=False))
    
    if 'detail' in report and not report['detail'].empty:
        if args.verbose:
            print("\n📝 ДЕТАЛИЗАЦИЯ ПО ЗАДАЧАМ:")
            print("="*100)
            pd.set_option('display.max_columns', None)
            pd.set_option('display.width', None)
            print(report['detail'].to_string(index=False))
    
    if 'issues' in report and not report['issues'].empty:
        print("\n⚠️ ПРОБЛЕМНЫЕ ЗАДАЧИ:")
        print("="*100)
        pd.set_option('display.max_columns', None)
        pd.set_option('display.width', None)
        print(report['issues'].to_string(index=False))
    
    print("\n" + "="*100)
    print(f"💰 ВСЕГО ПРОЕКТОВ: {report['total_projects']}")
    print(f"📦 ВСЕГО ЗАДАЧ:    {report['total_tasks']}")
    print(f"✅ КОРЕКТНЫХ:      {report['total_correct']}")
    print(f"⚠️  ПРОБЛЕМНЫХ:     {report['total_issues']}")
    print(f"⏱️  ВСЕГО ФАКТ:     {report['total_spent']:.2f} ч.")
    print(f"📏 ВСЕГО ОЦЕНКА:    {report['total_estimated']:.2f} ч.")
    print("="*100)
    
    if args.excel:
        filename = generate_excel(report)
        print(f"\n✅ Отчёт сохранён: {filename}")